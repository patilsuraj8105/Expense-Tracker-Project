from sqlalchemy.orm import Session
from app.models.expense import Expense
from datetime import date
from typing import Optional
import openpyxl
from io import BytesIO

def export_expenses_to_excel(db: Session, user_id: str, start_date: Optional[date], end_date: Optional[date]) -> BytesIO:
    query = db.query(Expense).filter(Expense.user_id == user_id)
    
    if start_date:
        query = query.filter(Expense.expense_date >= start_date)
    if end_date:
        query = query.filter(Expense.expense_date <= end_date)
        
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Headers
    headers = ["Title", "Amount (INR)", "Category", "Description", "Expense Date"]
    ws.append(headers)
    
    for expense in expenses:
        ws.append([
            expense.title,
            expense.amount,
            expense.category,
            expense.description or "",
            expense.expense_date.strftime("%Y-%m-%d") if expense.expense_date else ""
        ])
        # Format the amount cell (column 2) as currency (INR symbol)
        ws.cell(row=ws.max_row, column=2).number_format = '"₹"#,##0.00'
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
